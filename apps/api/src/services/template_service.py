"""Excel template generation service for Polish Ulga B+R project data."""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


def generate_excel_template() -> BytesIO:
    """
    Generate Excel template with Polish headers and validation rules.

    Returns:
        BytesIO: Buffer containing the generated Excel file.
    """
    wb = Workbook()

    # Create main data sheet
    ws_data = wb.active
    if ws_data is None:
        raise ValueError("Failed to create worksheet")
    ws_data.title = "Dane projektów"

    # Add Polish column headers with bold formatting
    headers = [
        "Nazwa projektu",
        "Opis",
        "Data rozpoczęcia",
        "Data zakończenia",
        "Cel projektu",
        "Osoba odpowiedzialna",
    ]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFC344", end_color="FFC344", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Add example data rows (Polish project examples)
    examples = [
        (
            "Opracowanie nowego systemu diagnostyki medycznej opartego na sztucznej inteligencji",
            "Projekt ma na celu stworzenie innowacyjnego systemu diagnostyki wykorzystującego sztuczną inteligencję "
            "do wczesnego wykrywania chorób neurodegeneracyjnych. System wykorzysta zaawansowane algorytmy uczenia "
            "maszynowego oraz przetwarzanie obrazów medycznych w celu zwiększenia dokładności diagnozy.",
            datetime(2024, 1, 15),
            datetime(2024, 12, 31),
            "Zwiększenie dokładności diagnostyki chorób neurodegeneracyjnych o 30% "
            "przy jednoczesnym skróceniu czasu badania do 15 minut",
            "Dr Jan Kowalski",
        ),
        (
            "Rozwój ekologicznego materiału opakowaniowego z recyklingu tworzyw sztucznych",
            "Badania nad opracowaniem nowego, w pełni biodegradowalnego materiału opakowaniowego powstającego "
            "z recyklingu odpadów plastikowych. Projekt obejmuje optymalizację procesu chemicznego degradacji "
            "oraz testowanie właściwości mechanicznych i barierowych nowego materiału.",
            datetime(2024, 3, 1),
            datetime(2025, 2, 28),
            "Stworzenie materiału opakowaniowego o wytrzymałości zbliżonej do standardowego plastiku, "
            "który ulega pełnej biodegradacji w ciągu 6 miesięcy",
            "Mgr inż. Anna Nowak",
        ),
        (
            "Automatyzacja procesów kontroli jakości w produkcji elektroniki z wykorzystaniem wizji komputerowej",
            "Projekt zakłada wdrożenie systemu automatycznej kontroli jakości wykorzystującego techniki wizji "
            "komputerowej i deep learning. System będzie wykrywał defekty produkcyjne w czasie rzeczywistym "
            "na linii produkcyjnej, eliminując konieczność manualnej inspekcji.",
            datetime(2024, 6, 1),
            datetime(2024, 11, 30),
            "Redukcja błędów produkcyjnych o 95% oraz skrócenie czasu kontroli jakości o 80% "
            "poprzez automatyzację procesów inspekcji",
            "Dr Piotr Wiśniewski",
        ),
    ]

    for row_idx, example in enumerate(examples, start=2):
        ws_data.cell(row=row_idx, column=1, value=example[0])
        ws_data.cell(row=row_idx, column=2, value=example[1])
        ws_data.cell(row=row_idx, column=3, value=example[2])
        ws_data.cell(row=row_idx, column=4, value=example[3])
        ws_data.cell(row=row_idx, column=5, value=example[4])
        ws_data.cell(row=row_idx, column=6, value=example[5])

        # Format date cells
        ws_data.cell(row=row_idx, column=3).number_format = "YYYY-MM-DD"
        ws_data.cell(row=row_idx, column=4).number_format = "YYYY-MM-DD"

    # Add comments with length requirements
    opis_comment = Comment(
        "Opis musi zawierać minimum 100 znaków. "
        "Opisz szczegółowo działania badawcze i rozwojowe projektu.",
        "System",
    )
    ws_data["B1"].comment = opis_comment

    cel_comment = Comment(
        "Cel projektu musi zawierać minimum 50 znaków. "
        "Określ konkretny, mierzalny cel badawczy projektu.",
        "System",
    )
    ws_data["E1"].comment = cel_comment

    # Add date validation for date columns
    date_validation = DataValidation(type="date", allow_blank=False)
    date_validation.error = "Wprowadź poprawną datę w formacie YYYY-MM-DD"
    date_validation.errorTitle = "Niepoprawna data"
    ws_data.add_data_validation(date_validation)

    # Apply validation to date columns (rows 2-21 for 1-20 projects)
    date_validation.add("C2:C21")
    date_validation.add("D2:D21")

    # Set column widths for readability
    ws_data.column_dimensions["A"].width = 60
    ws_data.column_dimensions["B"].width = 80
    ws_data.column_dimensions["C"].width = 18
    ws_data.column_dimensions["D"].width = 18
    ws_data.column_dimensions["E"].width = 60
    ws_data.column_dimensions["F"].width = 25

    # Create Instructions sheet
    ws_instructions = wb.create_sheet(title="Instrukcje")

    # Add instructions header
    ws_instructions["A1"] = "INSTRUKCJA WYPEŁNIANIA SZABLONU ULGA B+R"
    ws_instructions["A1"].font = Font(bold=True, size=14)

    # Add field descriptions
    instructions = [
        "WYMAGANE KOLUMNY:",
        "",
        "1. Nazwa projektu",
        "   - Pełna nazwa projektu badawczo-rozwojowego",
        "   - Przykład: 'Opracowanie nowego systemu diagnostyki medycznej'",
        "",
        "2. Opis",
        "   - Szczegółowy opis działań badawczych i rozwojowych",
        "   - WYMAGANE MINIMUM: 100 znaków",
        "   - Opisz metodologie, technologie, nowatorskie rozwiązania",
        "",
        "3. Data rozpoczęcia",
        "   - Data rozpoczęcia prac B+R w formacie YYYY-MM-DD",
        "   - Przykład: 2024-01-15",
        "",
        "4. Data zakończenia",
        "   - Data zakończenia prac B+R w formacie YYYY-MM-DD",
        "   - UWAGA: Musi być późniejsza lub równa dacie rozpoczęcia",
        "   - Przykład: 2024-12-31",
        "",
        "5. Cel projektu",
        "   - Konkretny, mierzalny cel badawczy",
        "   - WYMAGANE MINIMUM: 50 znaków",
        "   - Określ oczekiwane rezultaty i wskaźniki sukcesu",
        "",
        "6. Osoba odpowiedzialna",
        "   - Imię i nazwisko kierownika projektu",
        "   - Przykład: 'Dr Jan Kowalski'",
        "",
        "WAŻNE INFORMACJE:",
        "",
        "- W jednym pliku można umieścić od 1 do 20 projektów",
        "- Każdy wiersz reprezentuje jeden projekt",
        "- Przed wgraniem pliku sprawdź poprawność dat i długość opisów",
        "- System AI wygeneruje pełne karty projektów na podstawie tych danych",
        "",
        "Przykładowe dane znajdują się w arkuszu 'Dane projektów'.",
        "Możesz je usunąć i wpisać swoje projekty.",
    ]

    for row_idx, instruction_line in enumerate(instructions, start=2):
        cell = ws_instructions.cell(row=row_idx, column=1, value=instruction_line)
        if instruction_line.startswith("INSTRUKCJA") or instruction_line.startswith("WAŻNE"):
            cell.font = Font(bold=True)
        elif instruction_line and instruction_line[0].isdigit():
            cell.font = Font(bold=True, color="2C3E50")

    ws_instructions.column_dimensions["A"].width = 100

    # Save to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
