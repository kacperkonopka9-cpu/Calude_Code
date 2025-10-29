"""Integration tests for template download API endpoint."""

from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

from src.main import app


@pytest.mark.asyncio
async def test_download_template_endpoint() -> None:
    """Test GET /v1/templates/excel returns valid Excel file."""
    async with AsyncClient(app=app, base_url="http://test") as client:
        response = await client.get("/v1/templates/excel")

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "Szablon-Ulga-BR.xlsx" in response.headers["content-disposition"]


@pytest.mark.asyncio
async def test_downloaded_file_is_valid_excel() -> None:
    """Test that downloaded file can be opened as valid Excel workbook."""
    async with AsyncClient(app=app, base_url="http://test") as client:
        response = await client.get("/v1/templates/excel")

    # Verify it's a valid Excel file
    buffer = BytesIO(response.content)
    wb = load_workbook(buffer)

    assert len(wb.sheetnames) >= 1
    assert "Dane projektów" in wb.sheetnames


@pytest.mark.asyncio
async def test_health_check_endpoint() -> None:
    """Test that health check endpoint returns healthy status."""
    async with AsyncClient(app=app, base_url="http://test") as client:
        response = await client.get("/health")

    assert response.status_code == 200
    assert response.json() == {"status": "healthy"}
