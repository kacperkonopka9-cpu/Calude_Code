"""Unit tests for Excel template generation service."""

from io import BytesIO

from openpyxl import load_workbook

from src.services.template_service import generate_excel_template


def test_generate_template_returns_bytesio() -> None:
    """Test that generate_excel_template returns BytesIO buffer."""
    result = generate_excel_template()
    assert isinstance(result, BytesIO)


def test_template_has_correct_sheets() -> None:
    """Test that template has both required sheets."""
    buffer = generate_excel_template()
    wb = load_workbook(buffer)

    assert len(wb.sheetnames) == 2
    assert "Dane projektów" in wb.sheetnames
    assert "Instrukcje" in wb.sheetnames


def test_template_has_polish_headers() -> None:
    """Test that main sheet has all 6 required Polish column headers."""
    buffer = generate_excel_template()
    wb = load_workbook(buffer)
    ws = wb["Dane projektów"]

    expected_headers = [
        "Nazwa projektu",
        "Opis",
        "Data rozpoczęcia",
        "Data zakończenia",
        "Cel projektu",
        "Osoba odpowiedzialna",
    ]

    for col_idx, expected_header in enumerate(expected_headers, start=1):
        actual_header = ws.cell(row=1, column=col_idx).value
        assert actual_header == expected_header, f"Column {col_idx} header mismatch"


def test_template_has_example_rows() -> None:
    """Test that template contains 2-3 example data rows."""
    buffer = generate_excel_template()
    wb = load_workbook(buffer)
    ws = wb["Dane projektów"]

    # Check rows 2, 3, 4 have data (3 examples)
    for row_idx in range(2, 5):
        nazwa = ws.cell(row=row_idx, column=1).value
        opis = ws.cell(row=row_idx, column=2).value

        assert nazwa is not None, f"Row {row_idx} missing project name"
        assert opis is not None, f"Row {row_idx} missing description"
        assert len(str(opis)) >= 100, f"Row {row_idx} description too short"


def test_template_has_valid_example_data() -> None:
    """Test that example rows contain valid Polish project data."""
    buffer = generate_excel_template()
    wb = load_workbook(buffer)
    ws = wb["Dane projektów"]

    # Check first example row (row 2)
    nazwa = ws.cell(row=2, column=1).value
    opis = ws.cell(row=2, column=2).value
    data_rozpoczecia = ws.cell(row=2, column=3).value
    data_zakonczenia = ws.cell(row=2, column=4).value
    cel = ws.cell(row=2, column=5).value
    osoba = ws.cell(row=2, column=6).value

    assert isinstance(nazwa, str)
    assert isinstance(opis, str)
    assert len(opis) >= 100
    assert data_rozpoczecia is not None
    assert data_zakonczenia is not None
    assert isinstance(cel, str)
    assert len(cel) >= 50
    assert isinstance(osoba, str)


def test_template_date_validation() -> None:
    """Test that date columns have validation rules applied."""
    buffer = generate_excel_template()
    wb = load_workbook(buffer)
    ws = wb["Dane projektów"]

    # Check that data validation exists
    assert len(ws.data_validations.dataValidation) > 0, "No data validation found"


def test_template_file_size_under_500kb() -> None:
    """Test that template file size is ≤500KB."""
    buffer = generate_excel_template()
    file_size = len(buffer.getvalue())

    max_size = 500 * 1024  # 500KB in bytes
    assert file_size <= max_size, f"File size {file_size} exceeds 500KB limit"


def test_template_has_instructions_sheet() -> None:
    """Test that Instructions sheet contains Polish field descriptions."""
    buffer = generate_excel_template()
    wb = load_workbook(buffer)
    ws = wb["Instrukcje"]

    # Check header
    header = ws["A1"].value
    assert header is not None
    assert "INSTRUKCJA" in header or "Instrukcja" in header

    # Check that instructions mention required fields
    all_text = ""
    for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1):
        cell_value = row[0].value
        if cell_value:
            all_text += str(cell_value) + " "

    assert "Nazwa projektu" in all_text
    assert "Opis" in all_text
    assert "Data rozpoczęcia" in all_text
    assert "100 znaków" in all_text  # Minimum length requirement
